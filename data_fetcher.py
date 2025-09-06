# -*- coding: utf-8 -*-
import asyncio
import yaml
import os
import json
import openpyxl
import time
import http.server
import socketserver
import threading
import requests
from functools import partial
from datetime import datetime

# 导入 bilibili_fetcher 插件
def load_plugins(config):
    """加载必要的插件，特别是 bilibili_fetcher"""
    plugins = {}
    
    # 直接导入 bilibili_fetcher 插件
    try:
        import sys
        import os
        
        # 添加插件目录到系统路径
        plugins_dir = os.path.join('Video_Data_Analysis', 'Plugins')
        if not os.path.exists(plugins_dir):
            os.makedirs(plugins_dir, exist_ok=True)
            
        sys.path.append(plugins_dir)
        
        # 导入 bilibili_fetcher 插件
        from bilibili_fetcher import bilibili_fetcher
        
        # 获取插件配置
        plugin_config = config.get('plugins', {}).get('bilibili_fetcher', {})
        print(f"加载 bilibili_fetcher 插件，UID: {plugin_config.get('uid', '未设置')}")
        
        # 添加到插件列表
        plugins['bilibili_fetcher'] = bilibili_fetcher
        print("成功加载 bilibili_fetcher 插件")
    except Exception as e:
        print(f"加载 bilibili_fetcher 插件失败: {e}")
    
    return plugins

def format_large_number(num):
    """自动将大数字转换为万或亿为单位的字符串。"""
    if num >= 100000000:
        return f"{num / 100000000:.2f}亿"
    elif num >= 10000:
        return f"{num / 10000:.2f}万"
    else:
        return str(num)

def parse_refresh_interval(interval_str: str) -> int:
    unit = interval_str[-1].lower()
    value = int(interval_str[:-1])
    if unit == 's': return value
    if unit == 'm': return value * 60
    if unit == 'h': return value * 3600
    if unit == 'd': return value * 86400
    if unit == 'y': return value * 31536000
    return 86400

def load_excel_backup_for_dashboard(excel_dir):
    """从Excel备份文件读取数据用于仪表板"""
    try:
        backup_filepath = os.path.join(excel_dir, "bilibili_videos_latest.xlsx")
        if not os.path.exists(backup_filepath):
            print("未找到Excel备份文件")
            return None
        
        print(f"正在从Excel备份文件读取数据: {backup_filepath}")
        wb = openpyxl.load_workbook(backup_filepath)
        
        result = {'videos': [], 'follower_count': 'N/A'}
        
        # 读取汇总数据
        if "汇总数据" in wb.sheetnames:
            summary_ws = wb["汇总数据"]
            if summary_ws.max_row > 1:  # 有数据行
                row = list(summary_ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
                if row and len(row) > 0:
                    result['follower_count'] = row[0] if row[0] is not None else 'N/A'
        
        # 读取视频详情
        if "视频详情" in wb.sheetnames:
            videos_ws = wb["视频详情"]
            headers = [cell.value for cell in videos_ws[1]]
            
            for row in videos_ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:  # 确保有标题
                    video = {
                        'title': row[0] or '',
                        'link': row[1] or '',
                        'duration': row[2] or '',
                        'publish_time': row[3] or '',
                        'view_count': row[4] or 0,
                        'like_count': row[5] or 0,
                        'danmaku_count': row[6] or 0,
                        'comment_count': row[7] or 0,
                        'description': row[8] or '',
                        'fan_growth': '无法获取',
                        'danmakus': []  # 备份数据中不包含弹幕
                    }
                    result['videos'].append(video)
        
        wb.close()
        print(f"从Excel备份文件成功读取 {len(result['videos'])} 个视频数据")
        return result
        
    except Exception as e:
        print(f"从Excel备份文件读取数据时出错: {e}")
        return None

async def run_data_collection(config):
    output_dir = config.get('output', {}).get('directory', 'Excel')
    # 调整路径到正确的位置 - 现在是相对于根目录
    web_data_path = os.path.join('bilibili_analysis', 'Web', 'dashboard_data.json')
    # 确保目录存在
    os.makedirs(os.path.dirname(web_data_path), exist_ok=True)
    
    # 加载插件
    plugins = load_plugins(config)
    plugin_configs = config.get('plugins', {})
    
    # 用于存储所有插件返回的数据
    all_plugin_data = {}
    
    # 执行 bilibili_fetcher 插件获取数据
    if 'bilibili_fetcher' in plugins:
        try:
            plugin_config = plugin_configs.get('bilibili_fetcher', {})
            bilibili_fetcher = plugins['bilibili_fetcher']
            
            print(f"正在使用 bilibili_fetcher 插件获取数据...")
            bilibili_data = await bilibili_fetcher.fetch_user_videos(plugin_config, output_dir)
            
            if bilibili_data and 'videos' in bilibili_data:
                print(f"成功获取 {len(bilibili_data['videos'])} 个B站视频数据")
                all_plugin_data['bilibili_fetcher'] = bilibili_data
            else:
                print("未能获取有效的B站数据")
        except Exception as e:
            print(f"执行 bilibili_fetcher 插件时出错: {e}")
    else:
        print("未加载 bilibili_fetcher 插件，无法获取B站数据")

    summarize_data_for_dashboard(output_dir, web_data_path, all_plugin_data)

def summarize_data_for_dashboard(excel_dir, web_data_path, all_plugin_data):
    # 直接使用插件返回的数据，而不是从Excel文件重新读取
    print("开始处理仪表板数据...")
    
    # 从插件数据中获取视频数据和粉丝数
    bilibili_data = all_plugin_data.get('bilibili_fetcher', {})
    all_videos = bilibili_data.get('videos', [])
    follower_count = bilibili_data.get('follower_count', 'N/A')
    
    print(f"从插件获取到 {len(all_videos)} 个视频数据")

    # 如果获取到的视频数量为0，尝试从Excel备份文件读取数据
    if not all_videos:
        print("警告：未从插件获取到任何视频数据，尝试从Excel备份文件读取数据...")
        backup_data = load_excel_backup_for_dashboard(excel_dir)
        if backup_data and backup_data.get('videos'):
            all_videos = backup_data['videos']
            follower_count = backup_data.get('follower_count', follower_count)
            print(f"从Excel备份文件成功读取到 {len(all_videos)} 个视频数据")
        else:
            print("未能从Excel备份文件读取到有效数据，仪表板数据将不会更新。")
            return
    
    # 数据聚合
    total_views = sum(int(v.get('view_count', 0)) for v in all_videos)
    total_videos = len(all_videos)
    
    print(f"总视频数: {total_videos}")
    print(f"总播放量: {total_views}")

    # 按月份统计
    monthly_stats = {}
    for video in all_videos:
        try:
            pub_time = datetime.strptime(video['publish_time'], '%Y-%m-%d %H:%M:%S')
            month_key = pub_time.strftime('%Y-%m')
            if month_key not in monthly_stats:
                monthly_stats[month_key] = {'videos': 0, 'views': 0, 'likes': 0}
            monthly_stats[month_key]['videos'] += 1
            monthly_stats[month_key]['views'] += int(video.get('view_count', 0))
            monthly_stats[month_key]['likes'] += int(video.get('like_count', 0))
        except (ValueError, KeyError) as e:
            print(f"处理视频时间数据失败: {e}, 视频: {video.get('title', 'Unknown')}")
            continue

    sorted_months = sorted(monthly_stats.keys())

    # 计算月度变化
    for i, month_key in enumerate(sorted_months):
        stats = monthly_stats[month_key]
        if i > 0:
            prev_stats = monthly_stats[sorted_months[i-1]]
            
            # 播放量变化
            if prev_stats['views'] > 0:
                views_change = (stats['views'] - prev_stats['views']) / prev_stats['views'] * 100
                stats['views_change'] = f"{views_change:+.1f}%"
            else:
                stats['views_change'] = "N/A"
            
            # 视频数变化
            videos_change = stats['videos'] - prev_stats['videos']
            stats['videos_change'] = f"{videos_change:+}" if videos_change != 0 else "持平"

            # 点赞量变化
            if prev_stats['likes'] > 0:
                likes_change = (stats['likes'] - prev_stats['likes']) / prev_stats['likes'] * 100
                stats['likes_change'] = f"{likes_change:+.1f}%"
            else:
                stats['likes_change'] = "N/A"
        else:
            stats['views_change'] = "N/A"
            stats['videos_change'] = "N/A"
            stats['likes_change'] = "N/A"

    print(f"按月统计结果: {monthly_stats}")
    
    # 改进标签格式，包含年份信息
    trend_labels = []
    for m in sorted_months:
        year, month = m.split('-')
        trend_labels.append(f"{year}.{month}")
    
    # 准备累计和单月数据
    cumulative_videos, cumulative_views, cumulative_likes = 0, 0, 0
    trend_video_data, trend_views_data, trend_likes_data = [], [], []
    monthly_video_data, monthly_views_data, monthly_likes_data = [], [], []
    
    for month in sorted_months:
        # 累计数据
        cumulative_videos += monthly_stats[month]['videos']
        cumulative_views += monthly_stats[month]['views']
        cumulative_likes += monthly_stats[month]['likes']
        trend_video_data.append(cumulative_videos)
        trend_views_data.append(cumulative_views)
        trend_likes_data.append(cumulative_likes)
        
        # 单月数据
        monthly_video_data.append(monthly_stats[month]['videos'])
        monthly_views_data.append(monthly_stats[month]['views'])
        monthly_likes_data.append(monthly_stats[month]['likes'])

    # 计算近一个月数据和变化
    last_month_views = 0
    last_month_videos = 0
    last_month_likes = 0
    last_month_views_change = "N/A"
    last_month_videos_change = "N/A"
    last_month_likes_change = "N/A"

    if len(sorted_months) >= 1:
        last_month_key = sorted_months[-1]
        last_month_stats = monthly_stats[last_month_key]
        last_month_views = last_month_stats['views']
        last_month_videos = last_month_stats['videos']
        last_month_likes = last_month_stats['likes']

        # 计算相对于上个月的变化
        if len(sorted_months) >= 2:
            prev_month_key = sorted_months[-2]
            prev_month_stats = monthly_stats[prev_month_key]
            prev_month_views = prev_month_stats['views']
            prev_month_videos = prev_month_stats['videos']
            prev_month_likes = prev_month_stats['likes']

            # 播放量变化
            if prev_month_views > 0:
                views_change_percent = (last_month_views - prev_month_views) / prev_month_views * 100
                last_month_views_change = f"+{views_change_percent:.1f}%" if views_change_percent >= 0 else f"{views_change_percent:.1f}%"

            # 视频数量变化
            videos_change = last_month_videos - prev_month_videos
            if videos_change != 0:
                last_month_videos_change = f"+{videos_change}" if videos_change > 0 else str(videos_change)
            else:
                last_month_videos_change = "持平"

            # 点赞量变化
            if prev_month_likes > 0:
                likes_change_percent = (last_month_likes - prev_month_likes) / prev_month_likes * 100
                last_month_likes_change = f"+{likes_change_percent:.1f}%" if likes_change_percent >= 0 else f"{likes_change_percent:.1f}%"

    # 输出最终统计信息用于调试
    # print(f"\n=== 仪表板数据统计 ===")
    # print(f"总视频数: {total_videos}")
    # print(f"总播放量: {total_views:,}")
    # print(f"粉丝数: {follower_count}")
    # print(f"最近一个月播放量: {last_month_views:,}")
    # print(f"平均播放量: {total_views // total_videos if total_videos > 0 else 0:,}")
    # print(f"月份统计: {len(monthly_stats)} 个月")
    
    # 构建最终的JSON
    # 计算总点赞量
    total_likes = sum(int(v.get('like_count', 0)) for v in all_videos)
    print(f"总点赞量: {total_likes}")
    
    
    # 准备视频表现数据
    video_performance = []
    for v in all_videos:
        views = int(v.get('view_count', 0))
        if views >= 10000:
            likes = int(v.get('like_count', 0))
            video_performance.append({
                'title': v.get('title', '未知标题'),
                'views': views,
                'likes': likes,
                'like_rate': (likes / views) * 100 if views > 0 else 0
            })

    # 按点赞率排序
    video_performance.sort(key=lambda x: x['like_rate'], reverse=True)

    # 汇总所有弹幕，并附带视频标题
    all_danmakus = []
    for v in all_videos:
        if 'danmakus' in v and v['danmakus']:
            video_title = v.get('title', '未知视频')
            for dm_text in v['danmakus']:
                all_danmakus.append({'text': dm_text, 'video_title': video_title})

    dashboard_data = {
        "summary": {
            "total_fans": f"{follower_count:,}" if isinstance(follower_count, int) else follower_count,
            "total_views": format_large_number(total_views),
            "total_videos": total_videos,
            "total_likes": format_large_number(total_likes),
            "last_month_views": format_large_number(last_month_views),
            "last_month_views_change": last_month_views_change,
            "last_month_videos": last_month_videos,
            "last_month_videos_change": last_month_videos_change,
            "last_month_likes": format_large_number(last_month_likes),
            "last_month_likes_change": last_month_likes_change
        },
        "trend_chart": {
            "labels": trend_labels,
            "datasets": [
                { "label": "累计视频发布数", "data": trend_video_data },
                { "label": "累计播放量", "data": trend_views_data },
                { "label": "累计点赞量", "data": trend_likes_data },
                { "label": "每月视频发布数", "data": monthly_video_data },
                { "label": "每月播放量", "data": monthly_views_data },
                { "label": "每月点赞量", "data": monthly_likes_data }
            ]
        },
        "video_performance": video_performance,
        "monthly_details": monthly_stats,
        "follower_chart": {
            "labels": ["2022年", "2023年", "2024年", "2025年"],
            "datasets": [ {
                "label": "B站粉丝数量",
                "data": [
                    max(0, follower_count - 30) if isinstance(follower_count, int) else 0,
                    max(0, follower_count - 20) if isinstance(follower_count, int) else 0,
                    max(0, follower_count - 10) if isinstance(follower_count, int) else 0,
                    follower_count if isinstance(follower_count, int) else 0
                ]
            } ]
        },
        "additional_stats": {
            "videos_published_1": total_videos,
            "views_total_1": format_large_number(total_views),
            "videos_published_2": "N/A",
            "views_total_2": "N/A",
            "average_views": f"{total_views / total_videos:,.0f}" if total_videos > 0 else 0,
            "average_views_change_monthly": "N/A"
        },
        "all_videos": all_videos,
        "all_danmakus": all_danmakus
    }

    # 同时生成dashboard_data.json和bilibili_data.js
    # 1. 生成JSON文件（向后兼容）
    with open(web_data_path, 'w', encoding='utf-8') as f:
        json.dump(dashboard_data, f, ensure_ascii=False, indent=4)
        
    # 2. 生成JS文件，解决CORS问题
    bilibili_js_path = os.path.join(os.path.dirname(web_data_path), '..', '..', 'bilibili_data.js')
    with open(bilibili_js_path, 'w', encoding='utf-8') as f:
        f.write(f"window.BILIBILI_DATA = {json.dumps(dashboard_data, ensure_ascii=False, indent=2)};")
        
    print(f"已将仪表板数据写入到 '{web_data_path}' 和 '{bilibili_js_path}'")

def start_web_server(port: int):
    handler = partial(http.server.SimpleHTTPRequestHandler, directory='bilibili_analysis/Web')
    httpd = socketserver.TCPServer(("", port), handler)
    print(f"Web服务器已在 http://localhost:{port} 启动")
    httpd.serve_forever()

async def main():
    try:
        with open('config.yml', 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
    except FileNotFoundError:
        print("错误：找不到 config.yml 文件。")
        return
    except Exception as e:
        print(f"读取配置文件时出错: {e}")
        return

    server_config = config.get('server', {})
    port = server_config.get('port', 100)
    refresh_interval_str = server_config.get('refresh_interval', '1d')
    refresh_seconds = parse_refresh_interval(refresh_interval_str)

    web_config_path = os.path.join('bilibili_analysis', 'Web', 'web_config.json')
    with open(web_config_path, 'w', encoding='utf-8') as f:
        json.dump(config.get('web', {}), f, ensure_ascii=False, indent=4)

    server_thread = threading.Thread(target=start_web_server, args=(port,))
    server_thread.daemon = True
    server_thread.start()
    
    while True:
        print("开始执行数据刷新...")
        await run_data_collection(config)
        
        # 创建GitHub数据文件
        await generate_github_data(config)
        
        print(f"数据刷新完成。下一次刷新将在 {refresh_interval_str} 后进行。")
        await asyncio.sleep(refresh_seconds)

async def generate_github_data(config):
    """生成GitHub数据文件到项目根目录"""
    try:
        github_username = config.get('github', {}).get('username', 'Little100')
        print(f"正在获取GitHub用户 {github_username} 的数据...")
        
        # 设置GitHub数据的默认值
        github_data = {
            "projects": [],
            "languages": {
                "JavaScript": 45,
                "HTML": 25,
                "CSS": 15,
                "Python": 10,
                "Java": 5
            }
        }
        
        # 尝试通过API获取数据
        try:
            # 获取用户仓库信息
            repos_url = f"https://api.github.com/users/{github_username}/repos?sort=updated&per_page=100"
            headers = {'Accept': 'application/vnd.github.v3+json'}
            
            # 如果有GitHub token，使用它以提高API限制
            if 'token' in config.get('github', {}):
                headers['Authorization'] = f"token {config['github']['token']}"
                
            repos_response = requests.get(repos_url, headers=headers, timeout=10)
            repos_response.raise_for_status()
            repos = repos_response.json()
            
            # 处理仓库数据
            projects = []
            for repo in repos:
                if not repo.get('fork', False):  # 忽略fork的仓库
                    projects.append({
                        "name": repo.get('name', ''),
                        "fullName": repo.get('full_name', ''),
                        "description": repo.get('description', '') or '暂无描述',
                        "language": repo.get('language', 'Unknown'),
                        "stars": repo.get('stargazers_count', 0),
                        "forks": repo.get('forks_count', 0),
                        "url": repo.get('html_url', ''),
                        "updatedAt": repo.get('updated_at', ''),
                        "createdAt": repo.get('created_at', ''),
                        "topics": repo.get('topics', [])
                    })
            
            # 如果成功获取到仓库数据，则更新github_data
            if projects:
                github_data["projects"] = projects
                
                # 尝试获取语言统计
                languages_stats = {}
                for repo in repos[:20]:  # 限制为前20个仓库以避免API限制
                    try:
                        if 'languages_url' in repo and repo.get('size', 0) > 0:
                            lang_response = requests.get(repo['languages_url'], headers=headers, timeout=10)
                            lang_response.raise_for_status()
                            repo_langs = lang_response.json()
                            
                            for lang, bytes_count in repo_langs.items():
                                languages_stats[lang] = languages_stats.get(lang, 0) + bytes_count
                    except Exception as e:
                        print(f"获取仓库 {repo.get('name', '')} 的语言统计失败: {e}")
                
                # 计算语言百分比
                if languages_stats:
                    total_bytes = sum(languages_stats.values())
                    language_percentages = {}
                    
                    for lang, bytes_count in languages_stats.items():
                        percentage = round((bytes_count / total_bytes) * 100)
                        if percentage >= 1:  # 只包含占比至少1%的语言
                            language_percentages[lang] = percentage
                    
                    # 如果成功获取语言统计，更新github_data
                    if language_percentages:
                        github_data["languages"] = language_percentages
            
        except Exception as e:
            print(f"通过GitHub API获取数据失败: {e}")
            print("将使用默认的GitHub数据")
        
        # 将数据写入github_data.js文件（放在项目根目录）
        github_data_path = 'github_data.js'
        with open(github_data_path, 'w', encoding='utf-8') as f:
            f.write(f"window.GITHUB_DATA = {json.dumps(github_data, ensure_ascii=False, indent=2)};")
            
        print(f"已将GitHub数据写入到 '{github_data_path}'")
    except Exception as e:
        print(f"生成GitHub数据时出错: {e}")
        print("将使用默认的GitHub数据")
        
        # 即使出错，也确保生成一个默认的github_data.json文件
        github_data = {
            "projects": [],
            "languages": {
                "JavaScript": 45,
                "HTML": 25,
                "CSS": 15,
                "Python": 10,
                "Java": 5
            }
        }
        github_data_path = 'github_data.js'
        with open(github_data_path, 'w', encoding='utf-8') as f:
            f.write(f"window.GITHUB_DATA = {json.dumps(github_data, ensure_ascii=False, indent=2)};")

if __name__ == "__main__":
    try:
        # 尝试从根目录或Video_Data_Analysis目录加载配置
        config = None
        config_paths = ['config.yml', 'Video_Data_Analysis/config.yml']
        
        for config_path in config_paths:
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = yaml.safe_load(f)
                    print(f"成功从 {config_path} 加载配置")
                    break
            except FileNotFoundError:
                continue
        
        if config is None:
            print("错误：找不到配置文件。请确保存在 config.yml 或 Video_Data_Analysis/config.yml")
            exit(1)
            
        # 检查是否有必要的Python库
        try:
            import bilibili_api
            print("已安装 bilibili_api 库")
        except ImportError:
            print("警告: 未安装 bilibili_api 库，将无法获取B站数据")
            print("请使用以下命令安装: pip install bilibili-api-python")
            should_install = input("是否现在安装 bilibili-api-python? (y/n): ").lower().strip() == 'y'
            if should_install:
                import subprocess
                try:
                    subprocess.check_call([sys.executable, "-m", "pip", "install", "bilibili-api-python"])
                    print("已成功安装 bilibili-api-python")
                except Exception as e:
                    print(f"安装失败: {e}")
                    print("请手动安装: pip install bilibili-api-python")
        
        # 首次运行先生成GitHub数据
        print("开始生成GitHub数据...")
        asyncio.run(generate_github_data(config))
        
        # 创建B站数据
        print("开始生成B站数据...")
        asyncio.run(run_data_collection(config))
        
        print("\n数据生成完成！以下文件已创建：")
        print("1. github_data.json - 用于展示GitHub项目信息")
        print("2. bilibili_analysis/Web/dashboard_data.json - 用于展示B站数据分析")
        print("\n你现在可以通过以下命令启动HTTP服务器来查看你的个人主页：")
        print("python -m http.server 8000")
        print("然后在浏览器中访问 http://localhost:8000")
        
        # 可选：是否启动持续运行的服务器
        should_run_server = input("\n是否启动持续运行的数据更新服务？(y/n): ").lower().strip() == 'y'
        if should_run_server:
            # 然后启动主程序（每天更新一次数据）
            print("启动持续数据更新服务，每天会自动更新一次数据...")
            asyncio.run(main())
        else:
            print("脚本执行完毕。如需定期更新数据，请设置定时任务每天运行此脚本。")
    except KeyboardInterrupt:
        print("\n服务已停止。")
    except Exception as e:
        print(f"程序运行出错: {e}")
