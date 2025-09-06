import threading
import asyncio
from bilibili_api import user, video, sync
import openpyxl
from datetime import datetime
import os
import re

def save_to_excel(result, videos_by_year_month, output_dir, follower_count):
    """将视频数据保存到Excel文件"""
    try:
        # 确保输出目录存在
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        
        # 删除默认工作表
        wb.remove(wb.active)
        
        # 创建汇总工作表
        summary_ws = wb.create_sheet("汇总数据")
        summary_headers = ["粉丝数", "总视频数", "总播放量", "总点赞数", "总弹幕数", "总评论数", "更新时间"]
        summary_ws.append(summary_headers)
        
        total_videos = len(result['videos'])
        total_views = sum(int(v.get('view_count', 0)) for v in result['videos'])
        total_likes = sum(int(v.get('like_count', 0)) for v in result['videos'])
        total_danmakus = sum(int(v.get('danmaku_count', 0)) for v in result['videos'])
        total_comments = sum(int(v.get('comment_count', 0)) for v in result['videos'])
        
        summary_ws.append([
            follower_count,
            total_videos,
            total_views,
            total_likes,
            total_danmakus,
            total_comments,
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ])
        
        # 创建详细视频数据工作表
        videos_ws = wb.create_sheet("视频详情")
        video_headers = ["标题", "链接", "时长", "发布时间", "播放量", "点赞数", "弹幕数", "评论数", "简介"]
        videos_ws.append(video_headers)
        
        for video in result['videos']:
            videos_ws.append([
                video.get('title', ''),
                video.get('link', ''),
                video.get('duration', ''),
                video.get('publish_time', ''),
                video.get('view_count', 0),
                video.get('like_count', 0),
                video.get('danmaku_count', 0),
                video.get('comment_count', 0),
                video.get('description', '')[:500] if video.get('description') else ''  # 限制描述长度
            ])
        
        # 按年月创建工作表
        for year in sorted(videos_by_year_month.keys()):
            for month in sorted(videos_by_year_month[year].keys()):
                sheet_name = f"{year}年{month:02d}月"
                ws = wb.create_sheet(sheet_name)
                ws.append(video_headers)
                
                for video in videos_by_year_month[year][month]:
                    ws.append([
                        video.get('title', ''),
                        video.get('link', ''),
                        video.get('duration', ''),
                        video.get('publish_time', ''),
                        video.get('view_count', 0),
                        video.get('like_count', 0),
                        video.get('danmaku_count', 0),
                        video.get('comment_count', 0),
                        video.get('description', '')[:500] if video.get('description') else ''
                    ])
        
        # 保存Excel文件
        filename = f"bilibili_videos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(output_dir, filename)
        wb.save(filepath)
        print(f"视频数据已保存到Excel文件: {filepath}")
        
        # 同时保存一个最新的备份文件
        backup_filepath = os.path.join(output_dir, "bilibili_videos_latest.xlsx")
        wb.save(backup_filepath)
        print(f"最新备份已保存到: {backup_filepath}")
        
    except Exception as e:
        print(f"保存Excel文件时出错: {e}")

def load_from_excel_backup(output_dir):
    """从Excel备份文件读取数据"""
    try:
        backup_filepath = os.path.join(output_dir, "bilibili_videos_latest.xlsx")
        if not os.path.exists(backup_filepath):
            print("未找到Excel备份文件")
            return {'videos': [], 'follower_count': 'N/A'}
        
        print(f"正在从Excel备份文件读取数据: {backup_filepath}")
        wb = openpyxl.load_workbook(backup_filepath)
        
        result = {'videos': [], 'follower_count': 'N/A'}
        
        # 读取汇总数据
        if "汇总数据" in wb.sheetnames:
            summary_ws = wb["汇总数据"]
            if summary_ws.max_row > 1:  # 有数据行
                row = list(summary_ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
                if row and len(row) > 0:
                    result['follower_count'] = row[0] if row[0] is not None else 'N/A'
        
        # 读取视频详情
        if "视频详情" in wb.sheetnames:
            videos_ws = wb["视频详情"]
            headers = [cell.value for cell in videos_ws[1]]
            
            for row in videos_ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:  # 确保有标题
                    video = {
                        'title': row[0] or '',
                        'link': row[1] or '',
                        'duration': row[2] or '',
                        'publish_time': row[3] or '',
                        'view_count': row[4] or 0,
                        'like_count': row[5] or 0,
                        'danmaku_count': row[6] or 0,
                        'comment_count': row[7] or 0,
                        'description': row[8] or '',
                        'fan_growth': '无法获取',
                        'danmakus': []  # 备份数据中不包含弹幕
                    }
                    result['videos'].append(video)
        
        print(f"从Excel备份文件成功读取 {len(result['videos'])} 个视频数据")
        return result
        
    except Exception as e:
        print(f"从Excel备份文件读取数据时出错: {e}")
        return {'videos': [], 'follower_count': 'N/A'}

def get_suffix():
    return "_bilibili_fetcher"

def fetch_video_details(video_items, videos_by_year_month, successful_count, failed_count):
    for video_item in video_items:
        bvid = video_item['bvid']
        title = video_item.get('title', '未知标题')
        
        try:
            v = video.Video(bvid=bvid)
            detail = asyncio.run(v.get_info())
            
            pubdate = datetime.fromtimestamp(detail['pubdate'])
            year = pubdate.year
            month = pubdate.month
            
            if year not in videos_by_year_month:
                videos_by_year_month[year] = {}
            if month not in videos_by_year_month[year]:
                videos_by_year_month[year][month] = []
            
            view_count = detail.get('stat', {}).get('view', 0)
            if view_count == 0:
                view_count = detail.get('stat', {}).get('play', 0)
            
            danmakus = []
            try:
                dms = sync(v.get_danmakus(0))
                danmakus = [dm.text for dm in dms]
            except Exception as dm_e:
                print(f"获取视频 {bvid} 的弹幕失败: {dm_e}")

            video_details = {
                'title': detail['title'],
                'link': f"https://www.bilibili.com/video/{bvid}",
                'duration': f"{detail['duration'] // 60}:{detail['duration'] % 60:02d}",
                'publish_time': pubdate.strftime('%Y-%m-%d %H:%M:%S'),
                'view_count': view_count,
                'like_count': detail['stat']['like'],
                'danmaku_count': detail['stat']['danmaku'],
                'comment_count': detail['stat']['reply'],
                'description': detail['desc'],
                'fan_growth': '无法获取',
                'danmakus': danmakus
            }
            videos_by_year_month[year][month].append(video_details)
            successful_count[0] += 1

        except Exception as e:
            print(f"获取视频 {bvid} ({title}) 详情失败: {e}")
            failed_count[0] += 1

async def fetch_user_videos(plugin_config: dict, output_dir: str):
    uid = plugin_config.get('uid')
    if not uid:
        print("错误：在config.yml中未找到bilibili_fetcher插件的UID配置。")
        # 尝试从备份文件读取数据
        print("尝试从Excel备份文件读取数据...")
        return load_from_excel_backup(output_dir)

    try:
        u = user.User(uid)
        relation_info = await u.get_relation_info()
        follower_count = relation_info.get('follower', 0)

        page = 1
        all_videos = []
        while True:
            try:
                res = await u.get_videos(pn=page, ps=30)
                if not res.get('list', {}).get('vlist', []):
                    break
                all_videos.extend(res['list']['vlist'])
                page += 1
                await asyncio.sleep(0.5)
            except Exception as e:
                print(f"获取第 {page} 页视频时出错: {e}")
                break

        print(f"获取到 {len(all_videos)} 个视频的基本信息")
        
        # 如果没有获取到任何视频，尝试从备份文件读取
        if not all_videos:
            print("未获取到任何视频数据，尝试从Excel备份文件读取数据...")
            return load_from_excel_backup(output_dir)
        
        result = {
            'videos': [],
            'follower_count': follower_count
        }

        videos_by_year_month = {}
        successful_count = [0]
        failed_count = [0]

        # 将视频分割成8个部分
        num_threads = 8
        video_chunks = [all_videos[i::num_threads] for i in range(num_threads)]
        threads = []

        for chunk in video_chunks:
            thread = threading.Thread(target=fetch_video_details, args=(chunk, videos_by_year_month, successful_count, failed_count))
            threads.append(thread)
            thread.start()

        for thread in threads:
            thread.join()

        print(f"成功获取 {successful_count[0]} 个视频详情，失败 {failed_count[0]} 个")
        
        # 将 videos_by_year_month 中的视频数据添加到 result['videos']
        for year in sorted(videos_by_year_month.keys()):
            for month in sorted(videos_by_year_month[year].keys()):
                result['videos'].extend(videos_by_year_month[year][month])
        
        # 如果最终没有获取到视频详情，尝试从备份文件读取
        if not result['videos']:
            print("未获取到任何视频详情，尝试从Excel备份文件读取数据...")
            backup_result = load_from_excel_backup(output_dir)
            # 如果备份数据存在，使用备份数据但保留当前获取的粉丝数
            if backup_result['videos']:
                backup_result['follower_count'] = follower_count if follower_count else backup_result['follower_count']
                return backup_result
        
        # 保存到Excel文件
        if result['videos']:  # 只有在有数据时才保存
            save_to_excel(result, videos_by_year_month, output_dir, follower_count)

        return result

    except Exception as e:
        print(f"获取视频数据时出错: {e}")
        print("尝试从Excel备份文件读取数据...")
        return load_from_excel_backup(output_dir)

if __name__ == '__main__':
    test_uid = "1492647738" 
    test_output_dir = "Excel_test"
    asyncio.run(fetch_user_videos({'uid': test_uid}, test_output_dir))
